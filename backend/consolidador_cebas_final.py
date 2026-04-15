import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

print("--- ⚖️ CONSOLIDADOR OFICIAL CEBAS 2024 (100% PREENCHIDO) ---")

# DADOS UNIFICADOS (Extraídos de todas as suas planilhas: Ambulatório, Frequência e Paciente-Dia)
dados_2024 = [
    {"mes": "janeiro",   "data": "2024-01-01", "h_sus_qtd": 923,  "h_nsus_qtd": 93,  "h_sus_dia": 3502, "h_nsus_dia": 109, "a_sus": 2433, "a_nsus": 22},
    {"mes": "fevereiro", "data": "2024-02-01", "h_sus_qtd": 901,  "h_nsus_qtd": 112, "h_sus_dia": 3284, "h_nsus_dia": 134, "a_sus": 2350, "a_nsus": 87},
    {"mes": "março",     "data": "2024-03-01", "h_sus_qtd": 1013, "h_nsus_qtd": 101, "h_sus_dia": 3417, "h_nsus_dia": 117, "a_sus": 2419, "a_nsus": 75},
    {"mes": "abril",     "data": "2024-04-01", "h_sus_qtd": 980,  "h_nsus_qtd": 129, "h_sus_dia": 3466, "h_nsus_dia": 239, "a_sus": 2341, "a_nsus": 89},
    {"mes": "maio",      "data": "2024-05-01", "h_sus_qtd": 974,  "h_nsus_qtd": 100, "h_sus_dia": 3387, "h_nsus_dia": 141, "a_sus": 2385, "a_nsus": 74},
    {"mes": "junho",     "data": "2024-06-01", "h_sus_qtd": 934,  "h_nsus_qtd": 94,  "h_sus_dia": 3370, "h_nsus_dia": 124, "a_sus": 2379, "a_nsus": 88},
    {"mes": "julho",     "data": "2024-07-01", "h_sus_qtd": 909,  "h_nsus_qtd": 87,  "h_sus_dia": 3491, "h_nsus_dia": 121, "a_sus": 2470, "a_nsus": 78},
    {"mes": "agosto",    "data": "2024-08-01", "h_sus_qtd": 918,  "h_nsus_qtd": 124, "h_sus_dia": 3432, "h_nsus_dia": 132, "a_sus": 2429, "a_nsus": 112},
    {"mes": "setembro",  "data": "2024-09-01", "h_sus_qtd": 896,  "h_nsus_qtd": 99,  "h_sus_dia": 3133, "h_nsus_dia": 121, "a_sus": 2318, "a_nsus": 99},
    {"mes": "outubro",   "data": "2024-10-01", "h_sus_qtd": 999,  "h_nsus_qtd": 80,  "h_sus_dia": 3604, "h_nsus_dia": 98,  "a_sus": 2407, "a_nsus": 94},
    {"mes": "novembro",  "data": "2024-11-01", "h_sus_qtd": 1012, "h_nsus_qtd": 84,  "h_sus_dia": 3753, "h_nsus_dia": 94,  "a_sus": 2308, "a_nsus": 97},
    {"mes": "dezembro",  "data": "2024-12-01", "h_sus_qtd": 891,  "h_nsus_qtd": 100, "h_sus_dia": 3443, "h_nsus_dia": 102, "a_sus": 2238, "a_nsus": 88}
]

wb = Workbook()

# =========================================================
# PLAN 1 - APURAÇÃO DETALHADA
# =========================================================
ws1 = wb.active
ws1.title = "Plan1"

# Cabeçalho Oficial
ws1['A1'] = "HOSPITAL BENEFICENTE SANTA HELENA"
ws1['A3'] = "CEBAS - SAÚDE - Certificado de Entidades Beneficentes de Assistência Social - na área da saúde."
ws1['A5'] = "Competência de apresentação - Ano 2024"
ws1['A7'] = "DEMONSTRATIVO DAS INTERNAÇÕES E AMBULATÓRIO - 2024"

for r in [1, 3, 5, 7]:
    ws1[f'A{r}'].font = Font(bold=True, size=11)

# Estrutura Complexa de Colunas (Linhas 9 a 11)
ws1['A9'] = "Competência"
ws1['B9'] = "Internação"
ws1.merge_cells('B9:F9')
ws1['G9'] = "% SUS"
ws1['H9'] = "Ambulatório"
ws1.merge_cells('H9:J9')
ws1['K9'] = "% SUS MENSAL"

ws1['A10'] = "Mês/Ano"
ws1['B10'] = "Total - Pacientes"
ws1['C10'] = "SUS"
ws1.merge_cells('C10:D10')
ws1['E10'] = "NÃO SUS"
ws1.merge_cells('E10:F10')
ws1['G10'] = "Internação"
ws1['H10'] = "Total - Pacientes"
ws1['I10'] = "SUS"
ws1['J10'] = "NÃO SUS"
ws1['K10'] = ""

ws1['C11'] = "Quantidade"
ws1['D11'] = "Paciente-Dia"
ws1['E11'] = "Quantidade"
ws1['F11'] = "Paciente-Dia"
ws1['H11'] = ""
ws1['I11'] = "Quantidade"
ws1['J11'] = "Quantidade"

for row in ws1.iter_rows(min_row=9, max_row=11, min_col=1, max_col=11):
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Preenchendo os Dados (Linha 12 em diante)
linha = 12
for d in dados_2024:
    tot_int_qtd = d['h_sus_qtd'] + d['h_nsus_qtd']
    tot_amb_qtd = d['a_sus'] + d['a_nsus']
    
    ws1.cell(row=linha, column=1, value=d['data'])
    ws1.cell(row=linha, column=2, value=tot_int_qtd) 
    
    ws1.cell(row=linha, column=3, value=d['h_sus_qtd'])   # Quantidade SUS
    ws1.cell(row=linha, column=4, value=d['h_sus_dia'])   # Paciente-Dia SUS
    
    ws1.cell(row=linha, column=5, value=d['h_nsus_qtd'])  # Quantidade NÃO SUS
    ws1.cell(row=linha, column=6, value=d['h_nsus_dia'])  # Paciente-Dia NÃO SUS
    
    # Porcentagem de Internação baseada na Quantidade (como no seu modelo de 2023)
    c_int = ws1.cell(row=linha, column=7, value=(d['h_sus_qtd']/tot_int_qtd) if tot_int_qtd > 0 else 0)
    c_int.number_format = '0.00%'
    
    ws1.cell(row=linha, column=8, value=tot_amb_qtd)
    ws1.cell(row=linha, column=9, value=d['a_sus'])
    ws1.cell(row=linha, column=10, value=d['a_nsus'])
    
    # % SUS MENSAL GERAL
    tot_geral_sus = d['h_sus_qtd'] + d['a_sus']
    tot_geral = tot_int_qtd + tot_amb_qtd
    c_mensal = ws1.cell(row=linha, column=11, value=(tot_geral_sus/tot_geral) if tot_geral > 0 else 0)
    c_mensal.number_format = '0.00%'
    
    linha += 1

# Linhas de Soma (Aplicando fórmulas Excel dinâmicas)
ws1.cell(row=linha, column=1, value="SOMA").font = Font(bold=True)
ws1.cell(row=linha, column=2, value=f"=SUM(B12:B{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=3, value=f"=SUM(C12:C{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=4, value=f"=SUM(D12:D{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=5, value=f"=SUM(E12:E{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=6, value=f"=SUM(F12:F{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=7, value=f"=C{linha}/B{linha}").number_format = '0.00%'
ws1.cell(row=linha, column=7).font = Font(bold=True)
ws1.cell(row=linha, column=8, value=f"=SUM(H12:H{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=9, value=f"=SUM(I12:I{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=10, value=f"=SUM(J12:J{linha-1})").font = Font(bold=True)
ws1.cell(row=linha, column=11, value=f"=(C{linha}+I{linha})/(B{linha}+H{linha})").number_format = '0.00%'
ws1.cell(row=linha, column=11).font = Font(bold=True)

for col in ws1.columns:
    ws1.column_dimensions[col[0].column_letter].width = 16

# =========================================================
# PLAN 2 - RESUMO E SOMA GERAL
# =========================================================
ws2 = wb.create_sheet(title="Plan2")

ws2['B7'] = "HOSPITALAR"
ws2['H7'] = "AMBULATORIAL"
ws2['N7'] = "TOTAL ANUAL 2024"
for cell in ['B7', 'H7', 'N7']: ws2[cell].font = Font(bold=True)

headers_plan2 = [
    "Mês", "Total Hospitalar", "Frequência SUS", "Frequência Não SUS", "% Internação",
    "", "Mês", "Total Ambulatorial", "Qtd. Proc. SUS", "Qtd.Proc.Não SUS", "% Ambulatorio",
    "", "TOTAL GERAL", "Total SUS", "Total não SUS", "% SUS Mensal"
]
for col, h in enumerate(headers_plan2, start=2):
    c = ws2.cell(row=8, column=col, value=h)
    c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True, horizontal="center")

linha = 9
t = {'h_tot':0, 'h_sus':0, 'h_nsus':0, 'a_tot':0, 'a_sus':0, 'a_nsus':0, 'g_tot':0, 'g_sus':0, 'g_nsus':0}

for d in dados_2024:
    tot_hosp = d['h_sus_qtd'] + d['h_nsus_qtd'] 
    tot_amb = d['a_sus'] + d['a_nsus']
    tot_geral = tot_hosp + tot_amb
    tot_sus = d['h_sus_qtd'] + d['a_sus']
    tot_nsus = d['h_nsus_qtd'] + d['a_nsus']
    
    ws2.cell(row=linha, column=2, value=d['mes'])
    ws2.cell(row=linha, column=3, value=tot_hosp)
    ws2.cell(row=linha, column=4, value=d['h_sus_qtd'])
    ws2.cell(row=linha, column=5, value=d['h_nsus_qtd'])
    ws2.cell(row=linha, column=6, value=(d['h_sus_qtd']/tot_hosp) if tot_hosp>0 else 0).number_format = '0.00%'
    
    ws2.cell(row=linha, column=8, value=d['mes'])
    ws2.cell(row=linha, column=9, value=tot_amb)
    ws2.cell(row=linha, column=10, value=d['a_sus'])
    ws2.cell(row=linha, column=11, value=d['a_nsus'])
    ws2.cell(row=linha, column=12, value=(d['a_sus']/tot_amb) if tot_amb>0 else 0).number_format = '0.00%'
    
    ws2.cell(row=linha, column=14, value=tot_geral)
    ws2.cell(row=linha, column=15, value=tot_sus)
    ws2.cell(row=linha, column=16, value=tot_nsus)
    ws2.cell(row=linha, column=17, value=(tot_sus/tot_geral) if tot_geral>0 else 0).number_format = '0.00%'
    
    # Soma dos Acumulados
    t['h_tot']+=tot_hosp; t['h_sus']+=d['h_sus_qtd']; t['h_nsus']+=d['h_nsus_qtd']
    t['a_tot']+=tot_amb; t['a_sus']+=d['a_sus']; t['a_nsus']+=d['a_nsus']
    t['g_tot']+=tot_geral; t['g_sus']+=tot_sus; t['g_nsus']+=tot_nsus
    
    linha += 1

# Linha Final de SOMA Plan 2
ws2.cell(row=linha, column=2, value="SOMA").font = Font(bold=True)
ws2.cell(row=linha, column=3, value=t['h_tot']).font = Font(bold=True)
ws2.cell(row=linha, column=4, value=t['h_sus']).font = Font(bold=True)
ws2.cell(row=linha, column=5, value=t['h_nsus']).font = Font(bold=True)
ws2.cell(row=linha, column=6, value=(t['h_sus']/t['h_tot'])).number_format = '0.00%'
ws2.cell(row=linha, column=6).font = Font(bold=True)

ws2.cell(row=linha, column=9, value=t['a_tot']).font = Font(bold=True)
ws2.cell(row=linha, column=10, value=t['a_sus']).font = Font(bold=True)
ws2.cell(row=linha, column=11, value=t['a_nsus']).font = Font(bold=True)
ws2.cell(row=linha, column=12, value=(t['a_sus']/t['a_tot'])).number_format = '0.00%'
ws2.cell(row=linha, column=12).font = Font(bold=True)

ws2.cell(row=linha, column=14, value=t['g_tot']).font = Font(bold=True)
ws2.cell(row=linha, column=15, value=t['g_sus']).font = Font(bold=True)
ws2.cell(row=linha, column=16, value=t['g_nsus']).font = Font(bold=True)
ws2.cell(row=linha, column=17, value=(t['g_sus']/t['g_tot'])).number_format = '0.00%'
ws2.cell(row=linha, column=17).font = Font(bold=True)

for col in ws2.columns:
    ws2.column_dimensions[col[0].column_letter].width = 14

nome_saida = "Demonstrativo_CEBAS_Consolidado_2024.xlsx"
try:
    wb.save(nome_saida)
    print(f"🎉 SUCESSO! A planilha 2024 consolidada está pronta: {os.path.abspath(nome_saida)}")
    print("💡 Todas as colunas de Quantidade e Paciente-Dia foram integralmente preenchidas!")
except PermissionError:
    print(f"❌ ERRO: O arquivo '{nome_saida}' está ABERTO no seu Excel! Feche e rode novamente.")