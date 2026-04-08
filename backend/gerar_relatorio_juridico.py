import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import csv
import os

print("--- ⚖️ INICIANDO GERAÇÃO DO RELATÓRIO JURÍDICO (Padrão Ouro) ---")

# Nomes dos arquivos (Certifique-se de que estão na mesma pasta)
ARQUIVO_ANUAL = "Supabase Snippet Resumo anual de valores de AIHs (1).csv"
ARQUIVO_MENSAL = "Supabase Snippet Monthly SIH-SUS Financial Summary.csv"
ARQUIVO_JUROS = "JUROS.csv"
ARQUIVO_CORMOR = "CORMOR.csv"
ARQUIVO_SAIDA = "Consolidado_URV_Hospital_Santa_Helena_Completo.xlsx"

# Estilos idênticos ao Excel Jurídico
fonte_padrao = Font(name='Arial', size=10)
fonte_negrito = Font(name='Arial', size=10, bold=True)
borda_fina = Border(
    left=Side(style='thin', color='000000'), 
    right=Side(style='thin', color='000000'), 
    top=Side(style='thin', color='000000'), 
    bottom=Side(style='thin', color='000000')
)
centro = Alignment(horizontal='center', vertical='center')

def preencher_planilha(ws, df, tipo="anual"):
    # 1. Construir os Cabeçalhos
    cabecalhos = [
        ["01", "", "02", "03", "FATURAMENTO DO HOSPITAL CONSTANTE DOS ARQUIVOS", "", "", "", "", "09", "10", "11", "12", "13", "14"],
        ["Período", "", "Número de AIH's", "Obs.", "04", "05", "06", "07", "08", "DIFERENÇA 9,56%", "ÍNDICE DE ATUALIZAÇÃO", "VALOR DA DIFERENÇA ATUALIZADO", "JUROS CONFORME SENTENÇA", "VALOR DOS JUROS", "TOTAL GERAL"]
    ]
    
    if tipo == "anual":
        cabecalhos.append(["Linha", "Ano", "", "", "SERVIÇO HOSPITALAR", "SERV. PROFIS.", "SADT", "SANGUE", "SUBTOTAL", "", "", "", "", "", ""])
    else:
        cabecalhos.append(["Linha", "Competência", "", "", "SERVIÇO HOSPITALAR", "SERV. PROFIS.", "SADT", "SANGUE", "SUBTOTAL", "", "", "", "", "", ""])

    for linha in cabecalhos:
        ws.append(linha)

    # 2. Inserir os Dados Históricos
    linha_atual = 1
    total_sh = total_sp = total_sadt = total_sangue = total_sub = total_dif = 0

    for index, row in df.iterrows():
        # Corrige o problema do "2025.0" forçando para string e removendo o ".0"
        if tipo == "anual":
            periodo = str(row['Ano']).replace('.0', '')
        else:
            ano_limpo = str(row['Ano']).replace('.0', '')
            mes = str(row['Mes']).replace('.0', '').zfill(2)
            periodo = f"{mes}/{ano_limpo}"

        linha_dados = [
            linha_atual,                            # Linha
            periodo,                                # Ano ou MM/AAAA (Limpo)
            row['Numero_de_AIHs'],                  # Número de AIH's
            "",                                     # Obs
            row['Servico_Hospitalar'],              # 04 - SH
            row['Serv_Profissionais'],              # 05 - SP
            row['SADT'],                            # 06 - SADT
            row['SANGUE'],                          # 07 - SANGUE
            row['Subtotal'],                        # 08 - SUBTOTAL
            row['Diferenca_URV_9_56'],              # 09 - DIFERENÇA 9,56%
            "", "", "", "", ""                      # Vazio para o Jurídico calcular
        ]
        ws.append(linha_dados)
        
        # Acumular Totais
        total_sh += row['Servico_Hospitalar']
        total_sp += row['Serv_Profissionais']
        total_sadt += row['SADT']
        total_sangue += row['SANGUE']
        total_sub += row['Subtotal']
        total_dif += row['Diferenca_URV_9_56']
        
        linha_atual += 1

    # 3. Linha de Totais
    ws.append([]) # Linha em branco
    linha_totais = ["", "", "Valores Totais", "", total_sh, total_sp, total_sadt, total_sangue, total_sub, total_dif, "", "", "", "", ""]
    ws.append(linha_totais)
    ws.append([]) # Linha em branco

    # 4. Rodapé e Referências Legais
    rodape = [
        ["Fonte: arquivos rdmtaamm.dbc - Ministério da Saúde/DATASUS"],
        ["Observações:"],
        ["", "1 - Período de janeiro a junho não contemplado no processo"],
        ["", "2 - As internações relativas a competência 09/98 foram apresentadas ao DATASUS/MS juntamente à competência 10/98"],
        ["", "3 - Período de novembro a dezembro não contemplado no processo"],
        [],
        ["REFERÊNCIAS:"],
        ["01 =", "Ano/Mês do faturamento hospitalar apresentado, aprovado e pago pelo DATASUS/Ministério da Saúde"],
        ["02 =", "Número das Autorizações de Internação Hospitalar, origem dos valores pagos ao Hospital por internação."],
        ["03 =", "Código do Procedimento realizado de acordo com Tabela de Procedimentos DATASUS/MS, vigente à época."],
        ["04 =", "Valores referentes aos serviços realizados na Internação, devidos aos serviços do Hospital"],
        ["05 =", "Valores referentes aos serviços realizados na Internação, devidos aos profissionais médicos que atuaram."],
        ["06 =", "Valores referentes aos serviços de apoio diagnóstico e terapêutico (exames, fisioterapia e outros)."],
        ["07 =", "Valores referentes a transfusão de sangue e hemoderivados, quando necessário."],
        ["08 =", "Valor total dos serviços devidos ao Hospital pela internação, autorizados nas AIH's."],
        ["09 =", "Cálculo do valor da diferença de 9,56% originada na conversão de Cruzeiros Reais para Real."],
        ["10 =", "Índice de atualização de valores de acordo com a Tabela de Correção Monetária da Justiça Federal."],
        ["11 =", "Valor da diferença de 9,56% corrigida para valores atuais."],
        ["12 =", "Percentual dos juros a ser estabelecido em sentença"],
        ["13 =", "Valor dos juros estabelecidos em sentença"],
        ["14 =", "Total devido ao Hospital referente a soma dos valores constantes nas colunas 11 e 13"]
    ]

    for linha in rodape:
        ws.append(linha)

    # 5. Aplicar a Estética Rigorosa
    
    # Aplicar fonte Arial a TUDO
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.font = fonte_padrao

    # Cabeçalhos: Negrito, Centro e Bordas
    for r in range(1, 4):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.font = fonte_negrito
            cell.alignment = centro
            cell.border = borda_fina

    # Tabela de Dados: Bordas e Formato Moeda
    for r in range(4, linha_atual + 5): # Inclui a linha de Totais
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            
            # Aplica borda na grelha da tabela
            if r <= linha_atual + 4 and r != linha_atual + 3: 
                cell.border = borda_fina
                
            # Formato Moeda
            if c >= 5 and c <= 14:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'

    # Negrito na linha de Totais
    for c in range(1, 16):
        ws.cell(row=linha_atual+4, column=c).font = fonte_negrito

    # Ajustar largura das colunas para não cortar texto
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 32
    ws.column_dimensions['M'].width = 30
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 18

def anexar_aba_csv(wb, nome_arquivo, nome_aba):
    if os.path.exists(nome_arquivo):
        print(f"📄 A importar aba estática: {nome_aba}...")
        ws = wb.create_sheet(title=nome_aba)
        with open(nome_arquivo, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for row_data in reader:
                ws.append(row_data)
                
        # Estética básica para as abas anexadas
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.font = fonte_padrao
                
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    else:
        print(f"⚠️ Aviso: Ficheiro '{nome_arquivo}' não encontrado. Aba '{nome_aba}' ignorada.")

try:
    print("📂 A ler ficheiros CSV da base de dados...")
    df_anual = pd.read_csv(ARQUIVO_ANUAL)
    df_mensal = pd.read_csv(ARQUIVO_MENSAL)
    
    wb = openpyxl.Workbook()
    
    # Aba 1: Anual
    print("⚙️ A formatar Consolidado Anual...")
    ws_anual = wb.active
    ws_anual.title = "DIF 9,56% Anual"
    preencher_planilha(ws_anual, df_anual, tipo="anual")
    
    # Aba 2: Mensal
    print("⚙️ A formatar Detalhamento Mensal...")
    ws_mensal = wb.create_sheet(title="DIF 9,56% Mensal")
    preencher_planilha(ws_mensal, df_mensal, tipo="mensal")

    # Abas 3 e 4: Juros e Correção Monetária (Copiadas dos seus arquivos originais)
    anexar_aba_csv(wb, ARQUIVO_CORMOR, "CORMOR")
    anexar_aba_csv(wb, ARQUIVO_JUROS, "JUROS")

    wb.save(ARQUIVO_SAIDA)
    print(f"\n✅ SUCESSO ABSOLUTO! Ficheiro '{ARQUIVO_SAIDA}' gerado com perfeição estética.")

except FileNotFoundError as e:
    print(f"\n❌ ERRO: Ficheiro CSV principal não encontrado.\nDetalhe: {e}")
except Exception as e:
    print(f"\n❌ Ocorreu um erro inesperado: {e}")