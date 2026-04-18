import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import datetime

def formatar_relatorio_soulmv(arquivo_entrada, arquivo_saida):
    print("⏳ Lendo dados brutos e iniciando a formatação...")
    
    # 1. Lê os dados exportados da sua Query (Altere para read_csv se exportar como CSV)
    df = pd.read_excel(arquivo_entrada)

    # 2. Cria um novo arquivo Excel em branco para desenharmos
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Receita"

    # --- DEFINIÇÃO DE ESTILOS (O "CSS" do Excel) ---
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=11, color="FFFFFF") # Letra branca para o cabeçalho da tabela
    normal_font = Font(size=10)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Bordas finas padrão MV
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    # Fundo azul para a tabela e cinza para totais
    fill_blue = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # --- CABEÇALHO TIPO SOULMV (Topo da Página) ---
    ws.merge_cells('A1:I1')
    ws['A1'] = "HOSPITAL BENEFICENTE SANTA HELENA - CNES: 2311682"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:I2')
    ws['A2'] = "Relatório Analítico de Produção por Prestador SADT"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = align_center

    ws.merge_cells('A3:I3')
    data_emissao = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws['A3'] = f"Emitido em: {data_emissao}   |   Usuário: Coordenação Faturamento NII"
    ws['A3'].font = normal_font
    ws['A3'].alignment = align_center

    # Filtros e Parâmetros (Linhas 5 e 6)
    ws['A5'] = "Parâmetros do Relatório:"
    ws['A5'].font = bold_font
    ws['A6'] = "Competência: 01/2026"
    ws['C6'] = "Prestador: 489 - SANTA HELENA IMAGEM"
    ws['F6'] = "Filtro Ativo: REMOVIDO ECOCARDIOGRAFIA TRANSTORACICA (020501003)"
    ws['F6'].font = Font(color="FF0000", bold=True) # Destacado em vermelho para auditoria

    # --- CABEÇALHO DA TABELA ---
    # Estes devem ser exatamente os nomes que você quer mostrar na coluna
    colunas = ["Prestador", "Atendimento", "AIH", "Paciente", "Competência", "Cód. Procedimento", "Descrição do Exame", "Qtd", "Valor Líquido"]
    ws.append([""]) # Pula a linha 7
    ws.append(colunas) # Insere o cabeçalho na linha 8

    # Pinta o cabeçalho de Azul
    for cell in ws[8]:
        cell.font = header_font
        cell.fill = fill_blue
        cell.border = thin_border
        cell.alignment = align_center

    # --- PREENCHENDO OS DADOS E FORMATANDO ---
    row_idx = 9
    total_valor = 0

    for index, row in df.iterrows():
        # Mapeia as colunas do seu DataFrame (garanta que o nome bate com o SELECT da sua query)
        linha = [
            row.get('PRESTADOR', '489 - SANTA HELENA IMAGEM'),
            row.get('ATENDIMENTO', ''),
            row.get('AIH_PACIENTE', ''),
            row.get('NOME_PACIENTE', ''),
            row.get('PERIODO', '01/2026'),
            row.get('CODIGO_PROCEDIMENTO', ''),
            row.get('DESCRICAO_PROCEDIMENTO', ''),
            row.get('QTD', 0),
            row.get('VALOR_FATURADO', 0)
        ]
        ws.append(linha)
        
        # Somando o total financeiro
        total_valor += float(row.get('VALOR_FATURADO', 0))

        # Aplica bordas e alinhamentos célula por célula da linha atual
        for col_idx, celula in enumerate(ws[row_idx], 1):
            celula.border = thin_border
            celula.font = normal_font
            
            # Centraliza Atendimento, AIH, Competência, Código e QTD
            if col_idx in [2, 3, 5, 6, 8]: 
                celula.alignment = align_center
            # Formata a última coluna como Dinheiro (R$)
            elif col_idx == 9: 
                celula.number_format = 'R$ #,##0.00'
                
        row_idx += 1

    # --- LINHA DE TOTALIZADOR GERAL ---
    ws.append(["", "", "", "", "", "", "", "TOTAL DO REPASSE:", total_valor])
    total_row = row_idx
    
    # Formata a linha de Total
    for col_idx in range(1, 10):
        ws.cell(row=total_row, column=col_idx).fill = fill_gray
    
    ws.cell(row=total_row, column=8).font = bold_font
    ws.cell(row=total_row, column=8).alignment = align_right
    ws.cell(row=total_row, column=9).font = Font(bold=True, color="0066CC")
    ws.cell(row=total_row, column=9).number_format = 'R$ #,##0.00'
    ws.cell(row=total_row, column=9).border = thin_border

    # --- RODAPÉ DE SISTEMA ---
    ws.append([""])
    ws.append(["SOULMV - Sistema de Gestão Hospitalar", "", "", "", "", "", "", "", "Página 1 de 1"])
    ws.cell(row=total_row+2, column=1).font = Font(italic=True, size=8, color="808080")
    ws.cell(row=total_row+2, column=9).font = Font(italic=True, size=8, color="808080")
    ws.cell(row=total_row+2, column=9).alignment = align_right

    # --- AJUSTANDO A LARGURA DAS COLUNAS (Para nada ficar cortado) ---
    larguras = {'A': 28, 'B': 12, 'C': 16, 'D': 40, 'E': 13, 'F': 18, 'G': 45, 'H': 8, 'I': 16}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    # 3. Salva a Obra de Arte
    wb.save(arquivo_saida)
    print(f"✅ Sucesso! Relatório formatado salvo como: {arquivo_saida}")

# ==========================================
# COMO EXECUTAR:
# ==========================================
# 1. Salve o resultado da sua query como 'dados_brutos.xlsx' na mesma pasta.
# 2. Rode o script chamando a função abaixo:

if __name__ == "__main__":
    formatar_relatorio_soulmv('dados_brutos.xlsx', 'Relatorio_SADT_Oficial.xlsx')